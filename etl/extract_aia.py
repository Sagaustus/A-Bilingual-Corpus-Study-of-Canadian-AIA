"""
AIA Structured Extractor
========================
Reads every downloaded resource in resources/ and extracts a unified
structured record from each file, regardless of format (JSON, PDF, HTML,
CSV, XLSX).

Outputs:
  etl/output/extracted/{dataset_id}/{filename}.json  – per-file detail
  etl/output/aia_structured.csv                      – one row per file
  etl/output/extraction_log.csv                      – status per file
"""

import csv
import json
import re
import traceback
from pathlib import Path

import pdfplumber
import openpyxl
from bs4 import BeautifulSoup

RESOURCES_DIR = Path(__file__).parent.parent / "resources"
EXTRACT_DIR   = Path(__file__).parent / "output" / "extracted"
OUT_CSV       = Path(__file__).parent / "output" / "aia_structured.csv"
LOG_CSV       = Path(__file__).parent / "output" / "extraction_log.csv"

EXTRACT_DIR.mkdir(parents=True, exist_ok=True)

# ── AIA item-code decoder ─────────────────────────────────────────────────────
# JSON answers use codes like "item2-3" where the digit after "-" is the score.
# List answers like ["item3","item5"] are multi-select (we record count + values).

def score_from_item(val: str | list | None) -> int | None:
    """Extract the numeric score from a single item code, e.g. 'item2-3' → 3."""
    if isinstance(val, list):
        return None
    if not isinstance(val, str):
        return None
    m = re.search(r"-(\d+)$", val)
    return int(m.group(1)) if m else None


# AIA risk-profile dimension labels (riskProfile1 … riskProfile5)
RISK_LABELS = {
    "riskProfile1": "scope",
    "riskProfile2": "decision_impact",
    "riskProfile3": "automation_level",
    "riskProfile4": "data_sensitivity",
    "riskProfile5": "population_affected",
}

# Binary safeguard check sections → count Yes answers (score != 0)
SAFEGUARD_SECTIONS = {
    "consultationImplementation": "safeguard_consultation",
    "dataQualityImplementation":  "safeguard_data_quality",
    "fairnessImplementation":     "safeguard_fairness",
    "privacyImplementation":      "safeguard_privacy",
}


def detect_language(filename: str, text: str = "") -> str:
    name = filename.lower()
    if any(s in name for s in ["-fr", "_fr", "fra", "french", "francais", "français"]):
        return "fr"
    if any(s in name for s in ["-en", "_en", "eng", "english"]):
        return "en"
    if "bilingual" in name:
        return "bilingual"
    # Heuristic on text
    fr_words = len(re.findall(r"\b(le|la|les|de|du|des|et|en|ou|une|sur)\b", text[:2000], re.I))
    en_words = len(re.findall(r"\b(the|and|of|to|for|in|is|are|that|this)\b", text[:2000], re.I))
    if fr_words > en_words * 1.5:
        return "fr"
    if en_words > fr_words * 1.5:
        return "en"
    return "bilingual"


# ── JSON extractor ────────────────────────────────────────────────────────────

def extract_json(path: Path, dataset_id: str) -> dict:
    with open(path, encoding="utf-8") as f:
        raw = json.load(f)

    data   = raw.get("data", {})
    trans  = raw.get("translationsOnResult", {})

    # -- project details --
    rec = {
        "source_format":   "json",
        "version":         raw.get("version"),
        "language":        detect_language(path.name),
        "project_title_en": data.get("projectDetailsTitle") or trans.get("projectDetailsTitle"),
        "project_title_fr": trans.get("projectDetailsTitle") if trans.get("projectDetailsTitle") != data.get("projectDetailsTitle") else None,
        "department_code": data.get("projectDetailsDepartment-NS"),
        "branch":          data.get("projectDetailsBranch"),
        "respondent":      data.get("projectDetailsRespondent"),
        "phase":           data.get("projectDetailsPhase"),
        "phase_score":     score_from_item(data.get("projectDetailsPhase")),
        "description_en":  data.get("projectDetailsDescription"),
        "description_fr":  trans.get("projectDetailsDescription"),
        "program":         data.get("projectDetailsProgram"),
    }

    # -- risk profile scores --
    risk_total = 0
    for field, label in RISK_LABELS.items():
        s = score_from_item(data.get(field))
        rec[f"risk_{label}"] = s
        if s is not None:
            risk_total += s
    rec["risk_total_raw"] = risk_total

    # -- safeguard counts (number of Yes answers per section) --
    for prefix, col in SAFEGUARD_SECTIONS.items():
        yes_count = 0
        total     = 0
        for k, v in data.items():
            if k.startswith(prefix) and isinstance(v, str):
                s = score_from_item(v)
                if s is not None:
                    total += 1
                    if s != 0:
                        yes_count += 1
        rec[col]            = yes_count
        rec[col + "_total"] = total

    # -- system capabilities (multi-select list) --
    caps = data.get("aboutSystem1", [])
    rec["system_capabilities"] = ",".join(caps) if isinstance(caps, list) else caps
    rec["system_capabilities_count"] = len(caps) if isinstance(caps, list) else None

    # -- automation level (aboutAlgorithm1) --
    rec["algorithm_automation_level"] = score_from_item(data.get("aboutAlgorithm1"))
    rec["algorithm_uses_ml"]          = score_from_item(data.get("aboutAlgorithm2"))

    # -- decision sector (multi-select) --
    sector = data.get("decisionSector1", [])
    rec["decision_sectors"] = ",".join(sector) if isinstance(sector, list) else sector

    # -- impact fields (impact questions with scores) --
    impact_scores = {}
    for k, v in data.items():
        if k.startswith("impact") and re.match(r"impact\d+$", k):
            s = score_from_item(v)
            if s is not None:
                impact_scores[k] = s
    rec["impact_scores_json"] = json.dumps(impact_scores)
    rec["impact_score_sum"]   = sum(impact_scores.values())

    # -- store full raw data blob for deep analysis --
    rec["_raw_data_json"] = json.dumps(data, ensure_ascii=False)

    return rec


# ── PDF extractor ─────────────────────────────────────────────────────────────

# Patterns for key fields in the PDF text
_PDF_PATTERNS = {
    "impact_level":       r"Impact Level\s*[:\-–]\s*(\d+)",
    "current_score":      r"Current\s*Score\s*[:\-–]?\s*(\d+)",
    "raw_impact_score":   r"Raw\s*Impact\s*Score\s*[:\-–]?\s*(\d+)",
    "mitigation_score":   r"Mitigation\s*Score\s*[:\-–]?\s*(\d+)",
    "version":            r"Version\s*[:\-–]\s*([\d.]+)",
    "project_title_en":   r"(?:5\.\s*Project Title|5\.Project Title)[^\n]*\n(.+?)(?:\n\d+\.|\Z)",
    "department":         r"(?:3\.\s*Department|3\.Department)[^\n]*\n(.+?)(?:\n\d+\.|\Z)",
    "branch":             r"(?:4\.\s*Branch|4\.Branch)[^\n]*\n(.+?)(?:\n\d+\.|\Z)",
    "respondent":         r"(?:1\.\s*Name of Respondent|1\.Name\s*of\s*Respondent)[^\n]*\n(.+?)(?:\n\d+\.|\Z)",
    "phase":              r"(?:7\.|8\.)[^\n]*Phase[^\n]*\n(.+?)\s*\[\s*Points:\s*(\d+)\s*\]",
}

def extract_pdf(path: Path, dataset_id: str) -> dict:
    pages_text = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                pages_text.append(t)

    text = "\n".join(pages_text)
    lang = detect_language(path.name, text)

    rec: dict = {
        "source_format": "pdf",
        "language":      lang,
        "full_text_len": len(text),
        "page_count":    len(pages_text),
    }

    # Apply regex patterns
    for field, pattern in _PDF_PATTERNS.items():
        m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if m:
            if field == "phase":
                rec["phase"]       = m.group(1).strip()
                rec["phase_score"] = int(m.group(2))
            else:
                rec[field] = m.group(1).strip()[:500]

    # Cast numeric fields
    for num_field in ("impact_level", "current_score", "raw_impact_score", "mitigation_score"):
        if num_field in rec:
            try:
                rec[num_field] = int(rec[num_field])
            except ValueError:
                pass

    # Extract description (question 8 or 9)
    desc_m = re.search(
        r"(?:8\.|9\.)\s*Please provide a project description[:\s]*\n(.+?)(?:\n(?:About The System|Section \d))",
        text, re.IGNORECASE | re.DOTALL
    )
    if desc_m:
        rec["description"] = desc_m.group(1).strip()[:2000]

    # Extract Q&A section (Section 3)
    qa_m = re.search(r"Section 3[:\s]+Questions and Answers(.+?)(?:Section 4|\Z)",
                     text, re.IGNORECASE | re.DOTALL)
    if qa_m:
        rec["qa_section_text"] = qa_m.group(1).strip()[:5000]

    rec["_full_text"] = text

    return rec


# ── HTML extractor ────────────────────────────────────────────────────────────

def extract_html(path: Path, dataset_id: str) -> dict:
    with open(path, encoding="utf-8", errors="replace") as f:
        soup = BeautifulSoup(f, "html.parser")

    # Remove scripts/styles
    for tag in soup(["script", "style", "nav", "footer"]):
        tag.decompose()

    text = soup.get_text(separator="\n", strip=True)
    lang = detect_language(path.name, text)

    rec: dict = {
        "source_format": "html",
        "language":      lang,
        "full_text_len": len(text),
    }

    # Re-use PDF patterns on extracted text
    for field, pattern in _PDF_PATTERNS.items():
        m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        if m:
            if field == "phase":
                rec["phase"]       = m.group(1).strip()
                rec["phase_score"] = int(m.group(2))
            else:
                rec[field] = m.group(1).strip()[:500]

    rec["_full_text"] = text
    return rec


# ── CSV extractor ─────────────────────────────────────────────────────────────

def extract_csv(path: Path, dataset_id: str) -> dict:
    with open(path, encoding="utf-8-sig", errors="replace") as f:
        rows = list(csv.DictReader(f))

    return {
        "source_format": "csv",
        "language":      detect_language(path.name),
        "row_count":     len(rows),
        "columns":       ",".join(rows[0].keys()) if rows else "",
        "_raw_preview":  json.dumps(rows[:5], ensure_ascii=False),
    }


# ── XLSX extractor ────────────────────────────────────────────────────────────

def extract_xlsx(path: Path, dataset_id: str) -> dict:
    wb  = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(c is not None for c in row):
                rows.append([str(c) if c is not None else "" for c in row])
        sheets[sheet_name] = rows

    total_rows = sum(len(r) for r in sheets.values())
    return {
        "source_format": "xlsx",
        "language":      detect_language(path.name),
        "sheet_names":   ",".join(wb.sheetnames),
        "total_rows":    total_rows,
        "_raw_preview":  json.dumps(
            {k: v[:5] for k, v in sheets.items()}, ensure_ascii=False
        ),
    }


# ── dispatch ──────────────────────────────────────────────────────────────────

EXTRACTORS = {
    ".json": extract_json,
    ".pdf":  extract_pdf,
    ".html": extract_html,
    ".csv":  extract_csv,
    ".xlsx": extract_xlsx,
}


def extract_file(path: Path, dataset_id: str) -> dict | None:
    ext = path.suffix.lower()
    fn  = EXTRACTORS.get(ext)
    if fn is None:
        return None
    rec = fn(path, dataset_id)
    rec.update({
        "dataset_id":  dataset_id,
        "source_file": path.name,
    })
    return rec


# ── master CSV columns (flat, no _raw blobs) ─────────────────────────────────

MASTER_COLS = [
    "dataset_id", "source_file", "source_format", "language",
    "version",
    "project_title_en", "project_title_fr",
    "department", "department_code", "branch", "respondent",
    "phase", "phase_score",
    "program",
    "description", "description_en", "description_fr",
    # Risk profile
    "impact_level", "current_score", "raw_impact_score", "mitigation_score",
    "risk_total_raw",
    "risk_scope", "risk_decision_impact", "risk_automation_level",
    "risk_data_sensitivity", "risk_population_affected",
    # Safeguards
    "safeguard_consultation",      "safeguard_consultation_total",
    "safeguard_data_quality",      "safeguard_data_quality_total",
    "safeguard_fairness",          "safeguard_fairness_total",
    "safeguard_privacy",           "safeguard_privacy_total",
    # System info
    "system_capabilities", "system_capabilities_count",
    "algorithm_automation_level", "algorithm_uses_ml",
    "decision_sectors",
    "impact_score_sum",
    # CSV/XLSX extras
    "row_count", "columns", "sheet_names", "total_rows",
    "full_text_len", "page_count",
]


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    all_files = [
        p for p in RESOURCES_DIR.rglob("*")
        if p.is_file() and p.suffix.lower() in EXTRACTORS
    ]

    print(f"Extracting structured data from {len(all_files)} files...\n")

    master_rows: list[dict] = []
    log_rows:    list[dict] = []

    for path in sorted(all_files):
        dataset_id = path.parent.name
        rel        = path.relative_to(RESOURCES_DIR)

        try:
            rec = extract_file(path, dataset_id)
            if rec is None:
                continue

            # Save full per-file JSON (includes _raw blobs)
            out_dir = EXTRACT_DIR / dataset_id
            out_dir.mkdir(parents=True, exist_ok=True)
            out_path = out_dir / (path.stem + ".json")
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(rec, f, ensure_ascii=False, indent=2, default=str)

            # Flat row for master CSV (drop _private keys)
            flat = {k: rec.get(k, "") for k in MASTER_COLS}
            master_rows.append(flat)

            status = "ok"
            detail = ""
            icon   = "✓"

        except Exception:
            status = "error"
            detail = traceback.format_exc().splitlines()[-1]
            icon   = "✗"

        print(f"  {icon}  {rel}")
        log_rows.append({
            "file":       str(rel),
            "dataset_id": dataset_id,
            "format":     path.suffix.lower(),
            "status":     status,
            "detail":     detail,
        })

    # Write master CSV
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=MASTER_COLS, extrasaction="ignore")
        w.writeheader()
        w.writerows(master_rows)

    # Write log
    with open(LOG_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["file", "dataset_id", "format", "status", "detail"])
        w.writeheader()
        w.writerows(log_rows)

    ok    = sum(1 for r in log_rows if r["status"] == "ok")
    errs  = sum(1 for r in log_rows if r["status"] == "error")

    print(f"\nDone.  ✓ {ok} extracted  ✗ {errs} errors")
    print(f"Master CSV → {OUT_CSV}")
    print(f"Per-file JSON → {EXTRACT_DIR}/")


if __name__ == "__main__":
    main()
